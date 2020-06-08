from django.core.mail import EmailMultiAlternatives
from django.shortcuts import get_object_or_404, get_list_or_404
from django.utils.html import strip_tags
from django.template.loader import render_to_string
from django.utils.text import slugify
from functools import wraps
# from user.models import User
# from django.contrib.auth import get_user_model
from datetime import datetime
from django.http import HttpResponse, HttpResponseRedirect
from openpyxl import Workbook
from uuid import uuid4
from django.conf import settings
import re
import os

# User = get_user_model()



PRIVATE_IPS_PREFIX = ('10.', '192.', '172.',)


def get_client_ip(request):
    """ip kasi ke dare request mide ro begir"""
    remote_address = request.META.get('REMOTE_ADDR')
    ip = remote_address
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        proxies = x_forwarded_for.split(',')
        while (len(proxies) > 0 and proxies[0].startswith(PRIVATE_IPS_PREFIX)):
            proxies.pop(0)
        if len(proxies) > 0:
            ip = proxies[0]

    return ip


def send_email(*args, **kwargs):
    """send email from server to determined email
    
    Arguments:
        protocol {SCHEME} -- the protocol of the uri (HTTP/HTTPS)
        current_site {STRING} -- the current domain
    
    Keyword Arguments:
        the_messages {STRING} -- a message that should be in the body of email (default: {None})
        the_messages2 {STRING} -- a message that should be in the body of email (default: {None})
        is_team {STRING} -- specify if the email should be sent(for the team members of startup) (default: {None})
        first_name {STRING} -- the first name of the user that email should be sent (default: {None})
        last_name {STRING} -- the last name of the user that email should be sent (default: {None})
        sub {STRING} -- the subtitle of email (default: {None})
        template {STRING} -- the HTML file that contains email content (default: {None})
        to {STRING} -- the email address of the user that should be send to (default: {None})
        to2 {STRING} -- the email address of the user that should be send to (default: {None})
    
    Returns:
        BOOLEAN -- returns True if email has been sent successfuly
    """
    try:
        message = render_to_string(kwargs.get("template"), {
            'sub': kwargs.get('sub'),
            'messages': [message for message in kwargs.get("messages")],
            'domain': kwargs.get("domain") if kwargs.get("domain") else None,
            'to': kwargs.get("to")[0],
        })
        subject, from_email = kwargs.get('sub'), 'support@100startups.ir'
        # logger.info(f"email{message}---{to_the}")
        text_content = strip_tags(message)
        msg = EmailMultiAlternatives(
            subject, text_content, from_email, [client for client in kwargs.get('to')] )
        msg.attach_alternative(message, 'text/html')
        msg.send()
    except Exception as e:
        # logger.error(str(e))
        print(str(e))
    return True



def create_slug(instance):
    """automatic barassasse title slug misaze.
        age slug vase ye posti mojood bud id post ham mizare kenaresh ta unique bashe
    """
    # try:
    slug = slugify(("{}").format(instance.title), allow_unicode=True)
    # except:
    #     slug = slugify(("{}").format(instance.title), allow_unicode=True)
    qs = instance.__class__.objects.filter(slug=slug)
    if qs.exists():
        slug = ("{}-{}").format(slug, qs.first().id)
    return slug



def is_auth_or_not(function):
    @wraps(function)
    def wrapper(*args, user=None, **kwargs):
        request = kwargs['request']
        if request.user.is_authenticated:
            q = function(*args, user=request.user, **kwargs)
        elif request.user.is_anonymous:
            q = function(*args, user=user, **kwargs)
        return q
    return wrapper


def self_or_admin(request, Model, **instance):
    if request.user.is_staff:
        obj = get_object_or_404(Model, **instance)
    else:
        obj = get_object_or_404(Model, pk=request.user.pk)
    return obj


def search_engine(searches, instance):
    my_filter = {}
    results = None
    for search_field, search_val in searches.items():
        if search_val and search_val != "":
            field = search_field.split('___')
            if search_field in (f'{field[0]}___from_date', f'{field[0]}___to_date'):
                my_filter[f"{field[0]}__range"] = (searches.get(f'{field[0]}___from_date'), searches.get(f'{field[0]}___to_date'))
            elif search_val in ('True', 'False'):
                my_filter[f"{search_field}"] = search_val
            else:
                my_filter[f"{search_field}__icontains"] = search_val
            results = get_list_or_404(instance,**my_filter )
    return results


def export_excel(request, model, action=None):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-{movies}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
        movies=model,
    )
    workbook = Workbook()

    worksheet = workbook.active
    if model == 'user':
        if action:
            action = re.findall(r'\d+', action)
            model_qs = User.objects.filter(id__in=action)
        else:
            model_qs = User.objects.all()
        worksheet.title = 'کاربران'
        columns = ['ایدی','نام', 'نام خانوادگی', 'تلفن', 'جنسیت', 'نام کاربری', 'کد ملی', 'استان', 'شهر', 'تاریخ تولد', 'تحصیلات', 'مهارت', 'دنبال شده']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        
        for qs in model_qs:
            row_num += 1
            row = [qs.id, qs.first_name, qs.last_name, qs.phone, qs.sex, qs.username, qs.national_code, qs.profile.province, qs.profile.city, qs.profile.birth_date, qs.profile.grade,qs.profile.skill, ",".join([user.username for user in qs.profile.following.all()])]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        workbook.save(response)
    return response


def validate_image_extention(value):
    import os
    from django.core.exceptions import ValidationError
    ext = os.path.splitext(value.name)[-1]
    print(ext)
    valid_extentions = ['.jpg', '.png', '.jpeg']
    if ext.lower() not in valid_extentions:
        raise ValidationError("فایل مورد نظر پشتیبانی نمی شود.")


def hash_file_name(instance, filename):
    ext = filename.split('.')[-1]
    filename = f"{uuid4()}.{ext}"
    try:
        klass = instance._meta.model
    except:
        klass = instance.__class__.__name__
    return os.path.join(settings.MEDIA_ROOT, str(klass), '%Y-%m-%d', filename)

def logout_required(function):
    @wraps(function)
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated:
            return HttpResponseRedirect(request.META.get("HTTP_REFERER"))
        else:
            return function(request, *args, **kwargs)
    return wrapper